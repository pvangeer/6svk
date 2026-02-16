"""
Copyright (C) Stichting Deltares 2026. All rights reserved.

This file is part of the dikernel-python toolbox.

This program is free software; you can redistribute it and/or modify it under the terms of
the GNU Lesser General Public License as published by the Free Software Foundation; either
version 3 of the License, or (at your option) any later version.

This program is distributed in the hope that it will be useful, but WITHOUT ANY WARRANTY;
without even the implied warranty of MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.
See the GNU Lesser General Public License for more details.

You should have received a copy of the GNU Lesser General Public License along with this
program; if not, see <https://www.gnu.org/licenses/>.

All names, logos, and references to "Deltares" are registered trademarks of Stichting
Deltares and remain full property of Stichting Deltares at all times. All rights reserved.
"""

from svk.data import ResearchQuestion, Priority, ResearchLine, TimeFrame

from openpyxl import load_workbook
from pathlib import Path


class DatabaseReadError(Exception):
    """
    This class (Exception) can be raised by the research question database when it tries to read a database. It contains the row and column of the Excel file that could not be read as well as the message indicating why it cannot be read.
    """

    def __init__(self, message: str, i_row: int | None = None, i_column: int | None = None):
        super().__init__(message)
        self.i_row = i_row
        """The index of the row that could not be read/translated (zero based)."""
        self.i_column = i_column
        """The index of the column that could not be read/translated (zero based)."""

    @staticmethod
    def __number_to_letter(n: int) -> str:
        """
        Helper function that translates a column index (one based) to column name in Excel.

        :param n: The column index (one based, like in Excel)
        :type n: int
        :return: The Excel column name.
        :rtype: str
        """
        return chr(ord("A") + n - 1)

    @property
    def cell_reference(self) -> str:
        """
        Returns the cell reference of the Excel cell that could not be read.

        :param self: The DatabaseReadError
        :return: The cell that lead to the read error.
        :rtype: str
        """
        reference = ""
        if self.i_column is not None:
            reference = self.__number_to_letter(self.i_column + 1)

        if self.i_row is not None:
            return reference + str(self.i_row + 1)

        return reference


class Database(list[ResearchQuestion]):
    """
    Class that wraps a list[ResearchQuestion] to allow additional logic to read an convert a database file stored in Excel.
    """

    i_barrier = 0
    """Hard coded column number for the barrier"""
    i_id = 1
    """Hard coded column number for the question id"""
    i_reference_ids = 2
    """Hard coded column number for the references to other questions"""
    i_reference_question = 3
    """Hard coded column number for the reference number to the 160 questions list"""
    i_keywords = 4
    """Hard coded column number for the keywords"""
    i_question = 5
    """Hard coded column number for the question"""
    i_explanation = 6
    """Hard coded column number for the question explanation"""
    i_prio_water_safety = 7
    """Hard coded column number for the priority (water safety)"""
    i_prio_other_functions = 8
    """Hard coded column number for the priority (other functions)"""
    i_prio_management_maintenance = 9
    """Hard coded column number for the priority (management and maintenance)"""
    i_prio_operation = 10
    """Hard coded column number for the priority (operation)"""
    i_time_frame = 11
    """Hard coded column number for the time frame"""
    i_primary_research_line = 12
    """Hard coded column number for the primary research line"""
    i_secundary_research_line = 13
    """Hard coded column number for the secundary research line"""
    i_research_line_explanation = 14
    """Hard coded column number for the research line explanation"""
    i_status = 15
    """Hard coded column number for the status"""
    i_action_holder = 16
    """Hard coded column number for the action holder"""
    i_costs = 17
    """Hard coded column number for the costs"""
    i_lead_time = 18
    """Hard coded column number for the lead time"""

    def __init__(self, file_path: str):
        self.errors: list[DatabaseReadError] = []
        """A list of errors that can be filled during import/reading the database file."""

        if not self._check_file_path(file_path):
            raise ValueError("Excel file could not be found.")

        self.file_path: str = file_path
        """The file path of the Excel database file."""

    def _check_file_path(self, file_path: str) -> bool:
        """
        Check existance of the Excel file indicated as the database.

        :param self: The Database object
        :param file_path: file path that needs to be checked
        :type file_path: str
        :return: True in case the file exists, False if it doesn't exist or is not an Excel file.
        :rtype: bool
        """
        p = Path(file_path)
        return p.exists() and p.is_file() and p.suffix.lower() in {".xls", ".xlsx", ".xlsm", ".xlsb"}

    def read(self, sheet_name: str = "Database", first_data_row: int = 3):
        """
        Reads the database file.

        :param self: The Database object.
        :param sheet_name: The name of the Sheet that contains the database ("Database" by default)
        :type sheet_name: str
        :param first_data_row: The row number (1 based) of the first database record.
        :type first_data_row: int
        """
        wb = load_workbook(self.file_path)
        sheet = wb[sheet_name]

        for i_row, row in enumerate(sheet.iter_rows(min_row=first_data_row, max_row=None, values_only=True)):
            try:
                self.append(
                    ResearchQuestion(
                        id=Database._get_str(row, self.i_id),
                        question=Database._get_str(row, self.i_question),
                        explanation=Database._get_str_optional(row, self.i_explanation),
                        storm_surge_barrier=Database._get_str(row, self.i_barrier).split(","),  # TODO: Convert to StormSurgeBarrier
                        research_line_primary=Database._get_research_line_optional(row, self.i_primary_research_line),
                        research_line_secondary=Database._get_research_line_optional(row, self.i_secundary_research_line),
                        time_frame=Database._get_time_frame(row, self.i_time_frame),
                        prio_management_maintenance=Database._get_priority(row, self.i_prio_management_maintenance),
                        prio_other_functions=Database._get_priority(row, self.i_prio_other_functions),
                        prio_operation=Database._get_priority(row, self.i_prio_operation),
                        prio_water_safety=Database._get_priority(row, self.i_prio_water_safety),
                        action_holder=Database._get_str_optional(row, self.i_action_holder),
                        lead_time=Database._get_int_optional(row, self.i_lead_time),
                        costs_estimate=Database._get_int_optional(row, self.i_costs),
                        reference_ids=(
                            Database._get_str(row, self.i_reference_ids).replace(";", ",").split(",")
                            if not Database._empty(row, self.i_reference_ids)
                            else []
                        ),
                        reference_question=Database._get_int_optional(row, self.i_reference_question),
                    )
                )
            except DatabaseReadError as e:
                e.i_row = i_row + first_data_row - 1
                self.errors.append(e)
                continue

    @staticmethod
    def _get_str(row: tuple, i_column: int) -> str:
        value = row[i_column]
        if not isinstance(value, str):
            raise DatabaseReadError("Read cell is of incorrect type.", i_column=i_column)

        return str(value)

    @staticmethod
    def _get_str_optional(row: tuple, i_column: int) -> str | None:
        value = row[i_column]
        if not isinstance(value, str) or value == "":
            return None

        return str(value)

    @staticmethod
    def _get_int(row: tuple, i_column: int) -> int:
        value = row[i_column]
        if not isinstance(value, int):
            raise DatabaseReadError("Read cell is of incorrect type.", i_column=i_column)

        return int(value)

    @staticmethod
    def _get_int_optional(row: tuple, i_column: int) -> int | None:
        value = row[i_column]
        if not isinstance(value, int) or value is None:
            return None

        return int(value)

    @staticmethod
    def _get_research_line_optional(row: tuple, i_column: int) -> ResearchLine | None:
        value = row[i_column]
        if not isinstance(value, str) or value is None:
            return None

        return ResearchLine.get_research_line(int(str(value).split(".")[0]))

    @staticmethod
    def _empty(row: tuple, i_column: int) -> bool:
        value = row[i_column]
        return value == None

    @staticmethod
    def _get_priority(row: tuple, i_column: int) -> Priority:
        value = row[i_column]
        if not isinstance(value, int):
            return Priority.Unknown

        match int(value):
            case 1:
                return Priority.Low
            case 2:
                return Priority.Medium
            case 3:
                return Priority.High
            case _:
                return Priority.Unknown

    @staticmethod
    def _get_time_frame(row: tuple, i_column: int) -> TimeFrame:
        value = row[i_column]
        if not isinstance(value, int):
            return TimeFrame.Unknown

        match int(value):
            case 0:
                return TimeFrame.NotRelevant
            case 1:
                return TimeFrame.Now
            case 2:
                return TimeFrame.NearFuture
            case 3:
                return TimeFrame.Future
            case _:
                return TimeFrame.Unknown
