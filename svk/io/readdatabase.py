"""
Copyright (C) Stichting Deltares 2024. All rights reserved.

This file is part of the dikernel-python toolbox.

This program is free software; you can redistribute it and/or modify it under the terms of
the GNU Lesser General Public License as published by the Free Software Foundation; either
version 3 of the License, or (at your option) any later version.

This program is distributed in the hope that it will be useful, but WITHOUT ANY WARRANTY;
without even the implied warranty of MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.
See the GNU Lesser General Public License for more details.

You should have received a copy of the GNU Lesser General Public License along with this
program; if not, see <https://www.gnu.org/licenses/>.

All names, logos, and references to "Deltares" are registered trademarks of Stichting
Deltares and remain full property of Stichting Deltares at all times. All rights reserved.
"""

from svk.data import ResearchQuestion, Priority, ResearchLine, TimeFrame

from openpyxl import load_workbook


class DataBase(list[ResearchQuestion]):

    def __init__(self, file_path: str):
        if not self._check_file_path(file_path):
            raise ValueError("Excel file could not be found.")

        self.file_path: str = file_path

    def _check_file_path(self, file_path: str) -> bool:
        return True

    def read(self, sheet_name: str = "Database", first_data_row: int = 3) -> bool:
        wb = load_workbook(self.file_path)
        sheet = wb[sheet_name]

        for row in sheet.iter_rows(min_row=first_data_row, max_row=None, values_only=True):
            try:
                self.append(
                    ResearchQuestion(
                        question=DataBase._get_str(row[3]),
                        storm_surge_barrier=DataBase._get_str(row[0]).split(","),
                        action_holder=DataBase._get_str_optional(row[13]),
                        costs_estimate=DataBase._get_int_optional(row[14]),
                        lead_time=DataBase._get_int_optional(row[15]),
                        prio_budget=DataBase._get_priority(row[6]),
                        prio_functions=DataBase._get_priority(row[5]),
                        prio_operation=DataBase._get_priority(row[7]),
                        prio_water_safety=DataBase._get_priority(row[4]),
                        referece_number=[int(n) for n in DataBase._get_str(row[2]).split(",")],
                        reference_code=DataBase._get_str(row[1]).split(","),
                        research_line_primary=ResearchLine.get_research_line(DataBase._get_int(row[9])),
                        research_line_secondary=(
                            ResearchLine.get_research_line(DataBase._get_int(row[10]))
                            if DataBase._get_int_optional(row[10]) is not None
                            else None
                        ),
                        time_frame=DataBase._get_time_frame(row[8]),
                    )
                )
            except:
                ## TODO: Log row that could not be read.
                continue
        return True

    @staticmethod
    def _get_str(value) -> str:
        if not isinstance(value, str):
            raise ValueError("Read cell is of incorrect type.")

        return str(value)

    @staticmethod
    def _get_str_optional(value) -> str | None:
        if not isinstance(value, str) or value == "":
            return None

        return str(value)

    @staticmethod
    def _get_int(value) -> int:
        if not isinstance(value, int):
            raise ValueError("Read cell is of incorrect type.")

        return int(value)

    @staticmethod
    def _get_int_optional(value) -> int | None:
        if not isinstance(value, int) or value is None:
            return None

        return int(value)

    @staticmethod
    def _get_priority(value) -> Priority:
        if not isinstance(value, int):
            return Priority.Unknown

        match int(value):
            case 1:
                return Priority.Low
            case 2:
                return Priority.Medium
            case 3:
                return Priority.High
            case _:
                return Priority.Unknown

    @staticmethod
    def _get_time_frame(value) -> TimeFrame:
        if not isinstance(value, int):
            return TimeFrame.Unknown

        match int(value):
            case 1:
                return TimeFrame.Now
            case 2:
                return TimeFrame.NearFuture
            case 3:
                return TimeFrame.Future
            case _:
                return TimeFrame.Unknown
