from pathlib import Path
from openpyxl import load_workbook

from bisect import bisect_right
from typing import Any
from openpyxl.styles.fills import PatternFill
from svk.io._exceldatabase import DatabaseReadError
from svk.data import Driver, Function, Color, Grid, GridCell, GridHeader
from pydantic import BaseModel


class EndOfLifeCell(BaseModel):
    """
    Class that represents a cell in the EFL reader database. It is used to store the data in a structured way.
    """

    i_row: int
    i_column: int
    color: Color
    driver: Driver
    function: Function
    question_references: list[str] = []


class EndOfLifeDatabase:
    """
    Class that represents the EFL database. It is used to read the database file and store the data in a
    structured way.
    """

    def __init__(self, file_path: str):
        self.errors: list[DatabaseReadError] = []
        """A list of errors that can be filled during import/reading the database file."""

        self.row_header_categories_column: int | None = 1
        self.row_header_column: int = 2
        self.column_header_categories_row: int = 1
        self.column_header_row: int = 2
        self.first_data_column: int = 4
        self.first_data_row: int = 4
        self.columns_to_ignore: list[str] = ["Drivers"]

        self.sheet_name = "EFL"
        """Name of the sheet that contains the database. Default is 'Database'."""

        self.grid: Grid | None = None

        if not self._check_file_path(file_path):
            raise ValueError("Excel file could not be found.")

        self.file_path: str = file_path
        """The file path of the Excel database file."""

        self._drivers_dict: dict[int, int] = {}
        self._functions_dict: dict[int, int] = {}

    def _check_file_path(self, file_path: str) -> bool:
        """
        Check existance of the Excel file indicated as the database.

        :param self: The Database object
        :param file_path: file path that needs to be checked
        :type file_path: str
        :return: True in case the file exists, False if it doesn't exist or is not an Excel file.
        :rtype: bool
        """
        p = Path(file_path)
        return p.exists() and p.is_file() and p.suffix.lower() in {".xls", ".xlsx", ".xlsm", ".xlsb"}

    def read(self):
        """
        Reads the database file.

        :param self: The Database object.
        :param sheet_name: The name of the Sheet that contains the database ("Database" by default)
        :type sheet_name: str
        """
        wb = load_workbook(self.file_path, data_only=True)
        sheet = wb[self.sheet_name]

        _column_headers: dict[int, tuple[str, str]] = {}
        _row_headers: dict[int, tuple[str, str]] = {}

        _current_column_header_category: str = "Autonome situatie"
        for i_col in range(self.first_data_column, sheet.max_column + 1):
            header_category_cell = sheet.cell(row=self.column_header_categories_row, column=i_col)
            header_cell = sheet.cell(row=self.column_header_row, column=i_col)
            if header_category_cell.value is not None and isinstance(header_category_cell.value, str):
                _current_column_header_category = str(header_category_cell.value)

            if (
                header_cell.value is not None
                and isinstance(header_cell.value, str)
                and str(header_cell.value) not in self.columns_to_ignore
            ):
                _column_headers[i_col] = (_current_column_header_category, str(header_cell.value))

        _current_row_header_category: str = "Onbekend"
        for i_row in range(self.first_data_row, sheet.max_row + 1):
            if self.row_header_categories_column is not None:
                header_category_cell = sheet.cell(row=i_row, column=self.row_header_categories_column)
                if header_category_cell.value is not None and isinstance(header_category_cell.value, str):
                    _current_row_header_category = str(header_category_cell.value)

            header_cell = sheet.cell(row=i_row, column=self.row_header_column)
            if (
                header_cell.value is not None
                and isinstance(header_cell.value, str)
                and str(header_cell.value) not in self.columns_to_ignore
            ):
                _row_headers[i_row] = (_current_row_header_category, str(header_cell.value))

        _final_column_headers: list[GridHeader] = []
        _final_row_headers: list[GridHeader] = []
        _final_grid_cells: list[GridCell] = []
        is_first_column: bool = True
        for i_col_grid, i_col in enumerate(_column_headers.keys()):
            _final_column_headers.append(
                GridHeader(i_position=i_col_grid + 1, label=_column_headers[i_col][1], category=_column_headers[i_col][0])
            )

            for i_row_grid, i_row in enumerate(_row_headers.keys()):
                if is_first_column:
                    _final_row_headers.append(
                        GridHeader(i_position=i_row_grid + 1, label=_row_headers[i_row][1], category=_row_headers[i_row][0])
                    )
                _data_cell = sheet.cell(row=i_row, column=i_col)
                _final_grid_cells.append(
                    GridCell(
                        i_row=i_row_grid + 1,
                        i_column=i_col_grid + 1,
                        content=str(_data_cell.value),
                        color=self.fill_to_rgb(_data_cell.fill),
                    )
                )

            is_first_column = False

        self.grid = Grid(
            n_columns=len(_final_column_headers),
            n_rows=len(_final_row_headers),
            column_headers=_final_column_headers,
            row_headers=_final_row_headers,
            cells=_final_grid_cells,
        )

    @staticmethod
    def fill_to_rgb(fill: Any) -> Color:
        if not hasattr(fill, "fgColor"):
            return Color.White

        color = fill.fgColor

        rgb = "FFFFFF"
        if color.type == "rgb" and color.rgb:
            rgb = color.rgb[2:] if len(color.rgb) == 8 else color.rgb

        match rgb:
            case Color.White.value:
                return Color.White
            case Color.Yellow.value:
                return Color.Yellow
            case Color.Orange.value:
                return Color.Orange
            case Color.Red.value:
                return Color.Red
            case _:
                return Color.White
