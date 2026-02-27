from svk.data import (
    Priority,
    ResearchLine,
    TimeFrame,
    get_research_line,
    StormSurgeBarrier,
)

from pathlib import Path
from abc import ABC, abstractmethod
from openpyxl import load_workbook


class DatabaseReadError(Exception):
    """
    This class (Exception) can be raised by the research question database when it tries to read a database. It contains the row and column of the Excel file that could not be read as well as the message indicating why it cannot be read.
    """

    def __init__(self, message: str, i_row: int | None = None, i_column: int | None = None):
        super().__init__(message)
        self.i_row = i_row
        """The index of the row that could not be read/translated (zero based)."""
        self.i_column = i_column
        """The index of the column that could not be read/translated (zero based)."""

    @staticmethod
    def __number_to_letter(n: int) -> str:
        """
        Helper function that translates a column index (one based) to column name in Excel.

        :param n: The column index (one based, like in Excel)
        :type n: int
        :return: The Excel column name.
        :rtype: str
        """
        return chr(ord("A") + n - 1)

    @property
    def cell_reference(self) -> str:
        """
        Returns the cell reference of the Excel cell that could not be read.

        :param self: The DatabaseReadError
        :return: The cell that lead to the read error.
        :rtype: str
        """
        reference = ""
        if self.i_column is not None:
            reference = self.__number_to_letter(self.i_column + 1)

        if self.i_row is not None:
            return reference + str(self.i_row + 1)

        return reference


class ExcelDatabase(ABC):
    def __init__(self, file_path: str):
        self.errors: list[DatabaseReadError] = []
        """A list of errors that can be filled during import/reading the database file."""

        if not self._check_file_path(file_path):
            raise ValueError("Excel file could not be found.")

        self.file_path: str = file_path
        """The file path of the Excel database file."""

    def _check_file_path(self, file_path: str) -> bool:
        """
        Check existance of the Excel file indicated as the database.

        :param self: The Database object
        :param file_path: file path that needs to be checked
        :type file_path: str
        :return: True in case the file exists, False if it doesn't exist or is not an Excel file.
        :rtype: bool
        """
        p = Path(file_path)
        return p.exists() and p.is_file() and p.suffix.lower() in {".xls", ".xlsx", ".xlsm", ".xlsb"}

    def read(self, sheet_name: str = "Database", first_data_row: int = 3):
        """
        Reads the database file.

        :param self: The Database object.
        :param sheet_name: The name of the Sheet that contains the database ("Database" by default)
        :type sheet_name: str
        :param first_data_row: The row number (1 based) of the first database record.
        :type first_data_row: int
        """
        wb = load_workbook(self.file_path, data_only=True)
        sheet = wb[sheet_name]

        for i_row, row in enumerate(sheet.iter_rows(min_row=first_data_row, max_row=None, values_only=True)):
            try:
                self.read_and_append_row(row)
            except DatabaseReadError as e:
                e.i_row = i_row + first_data_row - 1
                self.errors.append(e)
                continue

    @abstractmethod
    def read_and_append_row(self, row):
        pass

    @staticmethod
    def _get_as_str(row: tuple, i_column: int) -> str:
        return str(row[i_column])

    @staticmethod
    def _get_str_optional(row: tuple, i_column: int) -> str | None:
        value = row[i_column]
        if not isinstance(value, str) or value == "":
            return None

        return str(value)

    @staticmethod
    def _get_int(row: tuple, i_column: int) -> int:
        value = row[i_column]
        if not isinstance(value, int):
            raise DatabaseReadError("Read cell is of incorrect type.", i_column=i_column)

        return int(value)

    @staticmethod
    def _get_int_optional(row: tuple, i_column: int) -> int | None:
        value = row[i_column]
        if not isinstance(value, int) or value is None:
            return None

        return int(value)

    @staticmethod
    def _get_research_line_optional(row: tuple, i_column: int) -> ResearchLine | None:
        value = row[i_column]
        if not isinstance(value, str) or value is None:
            return None

        return get_research_line(int(str(value).split(".")[0]))

    @staticmethod
    def _empty(row: tuple, i_column: int) -> bool:
        value = row[i_column]
        return value == None

    @staticmethod
    def _get_priority(row: tuple, i_column: int) -> Priority:
        value = row[i_column]
        if not isinstance(value, int):
            return Priority.Unknown

        match int(value):
            case 0:
                return Priority.No
            case 1:
                return Priority.Low
            case 2:
                return Priority.Medium
            case 3:
                return Priority.High
            case _:
                return Priority.Unknown

    @staticmethod
    def _get_time_frame(row: tuple, i_column: int) -> TimeFrame:
        value = row[i_column]
        if not isinstance(value, int):
            return TimeFrame.Unknown

        match int(value):
            case 0:
                return TimeFrame.NotRelevant
            case 1:
                return TimeFrame.Now
            case 2:
                return TimeFrame.NearFuture
            case 3:
                return TimeFrame.Future
            case _:
                return TimeFrame.Unknown

    @staticmethod
    def _get_storm_surge_barriers(row: tuple, i_column: int) -> list[StormSurgeBarrier]:
        barrier_strings = ExcelDatabase._get_as_str(row, i_column).split(",")
        barriers = []
        for b in barrier_strings:
            match b:
                case "HV":
                    barriers.append(StormSurgeBarrier.HaringvlietBarrier)
                case "HIJK":
                    barriers.append(StormSurgeBarrier.HollandseIJsselBarrier)
                case "6SVK" | "6SSB":
                    barriers.append(StormSurgeBarrier.All)
                case "OSK" | "ESB":
                    barriers.append(StormSurgeBarrier.EasternScheldBarrier)
                case "MLK" | "MLB":
                    barriers.append(StormSurgeBarrier.MaeslantBarrier)
                case "HK":
                    barriers.append(StormSurgeBarrier.HartelBarrier)
                case "RP":
                    barriers.append(StormSurgeBarrier.Ramspol)
                case _:
                    raise DatabaseReadError(
                        "Cannot read storm surge barrier. Should be on of ['6SVK','HV','HIJK','HK','MLK','OSK','RP'].", i_column=i_column
                    )
        return barriers
